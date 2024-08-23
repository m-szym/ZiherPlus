# Record loading and preprocessing for ZiherPlus
# Author: Marek Szymański

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import datetime
from typing import Generator, Optional


from zih_types import ZihRecord
from site_specific import FormFieldIDs
from excel_specific import COLS


def print_record(record: ZihRecord, nr: Optional[int]) -> None:
    '''Present the `record` in an easy to read way
    
    :param record: dict conatining input_ids, respective input values
                   and other info that's not sent to the form
    :param nr: index of the record in current context
    '''
    print(f"=== {nr if nr is not None else '#'} RECORD: {record['IDX']} ===")
    for k, v in record.items():
        print(k + ": " + str(v))
    print()

def iter_data(
    ws: Worksheet, min_row: int, max_row: int
) -> Generator[ZihRecord, None, None]:
    '''Iterates over rows in worksheet `ws` from `min_row` to `max_row`
    
    :param ws: Excel worksheet to iterate over
    :param min_row: first row of record data
    :param max_row: last  row of record data

    :returns: yields dicts containing input_ids, respective input values
              and other info about each record from each row between `min_row` and `max_row`
              {
                "type": record typ - either "income" or "cost",
                "category": category of the record - as in COLS,
                "IDX": index of the record in the Excel file
                "date": record date, in Y-m-d format
                id of form input, specific to record category and type: monetary value (income or cost)
                ...
              }
    '''
    for row in ws.iter_rows(min_row=min_row, max_row=max_row):
        entry = {}
        for k, v in COLS.items():
            k -= 1
            if not row[k].value:
                continue

            if v["type"] == "misc":
                if v["name"] == "data":
                    if (isinstance(row[k].value, datetime.datetime)):
                        entry[v["id"]] = str(row[k].value.strftime("%Y-%m-%d"))
                    else:
                        entry[v["id"]] = row[k].value
                else:
                    entry[v["id"]] = row[k].value
            elif v["type"] in ["income", "cost"]:
                entry[FormFieldIDs['amountFun'](v['id'])] = row[k].value
                entry["type"] = v["type"]
                entry["category"] = v["name"]

        yield entry

def load_workbook(
    filename: str, sheetname: Optional[str] = None
) -> tuple[Workbook, Worksheet]:
    '''Loads Excel workbook pointed to by `filename` and from it either the worksheet `sheetname` or the active worksheet
    
    :param filename: path to Excel file to load
    :param sheetname: sheetname to load from the workbook, leave out to load active worksheet

    :returns: tuple of loaded workbook and loaded worksheet
    '''
    wb = openpyxl.load_workbook(filename)
    ws = wb[sheetname] if sheetname else wb.active
    return wb, ws

def load_worksheet(wb: Workbook, sheetname: str) -> Worksheet:
    '''Wrapper for loading worksheet `sheetname` from workbook `wb`
    
    :param wb: Excel workbook from which to lead the worksheet
    :param sheetname: name of the worksheet to load

    :returns: loaded worksheet `sheetname`
    '''
    return wb[sheetname]
